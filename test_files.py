import os
import pytest

from zipfile import ZipFile
import csv
from openpyxl import load_workbook
import PyPDF2
from io import StringIO

script_dir = os.path.dirname(os.path.abspath(__file__))
resources_dir = os.path.join(script_dir, "resources")

# test data: filetype -> [filename, pattern_in_file]
data = {'xlsx': ['test.xlsx', 'glad'], 'csv': ['test.csv', 'miracle'], 'pdf': ['test.pdf', 'friend']}

xlsx_name, xlsx_pattern = data['xlsx']
csv_name, csv_pattern = data['csv']
pdf_name, pdf_pattern = data['pdf']
zip_name = 'test.zip'

file_xlsx = os.path.join(resources_dir, xlsx_name)
file_pdf = os.path.join(resources_dir, pdf_name)
file_csv = os.path.join(resources_dir, csv_name)

file_zip = os.path.join(resources_dir, zip_name)


def test_create_zip_file():

    if os.path.exists(file_zip):
        os.remove(file_zip)

    with ZipFile(file_zip, 'w') as zipfile:
        zipfile.write(file_xlsx, arcname=xlsx_name)
        zipfile.write(file_csv, arcname=csv_name)
        zipfile.write(file_pdf, arcname=pdf_name)


def test_verify_csv_file():

    with ZipFile(file_zip) as zip_handler:
        with zip_handler.open(csv_name, 'r') as csv_file:
            content_to_iterate = StringIO(csv_file.read().decode('latin-1'))
            content = csv.reader(content_to_iterate, delimiter=',')
            exists_in_csv = False

            for row in content:
                if csv_pattern in row:
                    exists_in_csv = True
                    break

            assert exists_in_csv


def test_verify_xlsx_file():

    with ZipFile(file_zip) as zip_handler:
        with zip_handler.open(xlsx_name, 'r') as xlsx_file:
            wb = load_workbook(filename=xlsx_file)
            sheet = wb.active
            exists_in_xlsx = False

            for r in range(1, sheet.max_row + 1):
                for c in range(1, sheet.max_column + 1):
                    cell = sheet.cell(r, c).value
                    if cell is not None and xlsx_pattern in cell:
                        exists_in_xlsx = True
                        break
                if exists_in_xlsx:
                    break

            assert exists_in_xlsx


def test_verify_pdf_file():

    with ZipFile(file_zip) as zip_handler:
        with zip_handler.open(pdf_name, 'r') as pdf_file:
            pdf = PyPDF2.PdfReader(pdf_file)
            num_pages = len(pdf.pages)
            exists_in_pdf = False

            for i in range(0, num_pages):
                page = pdf.pages[i]
                text = page.extract_text()
                if pdf_pattern in text:
                    exists_in_pdf = True
                    break

            assert exists_in_pdf
